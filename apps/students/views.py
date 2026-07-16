"""Student views."""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Q
from .models import Student
from .forms import StudentForm
from apps.classes.models import Classroom
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

@method_decorator(login_required, name='dispatch')
class StudentListView(View):
    template_name = 'students/list.html'

    def get(self, request):
        # Base Query
        students = Student.objects.select_related('classroom', 'user').all()
        
        # 1. Fitur Pencarian (Search)
        q = request.GET.get('q')
        if q:
            students = students.filter(
                Q(full_name__icontains=q) | Q(nis__icontains=q)
            )
            
        # 2. Fitur Filter Kelas (Diterapkan bersamaan dengan search)
        classroom_id = request.GET.get('classroom')
        if classroom_id:
            students = students.filter(classroom_id=classroom_id)

        context = {
            'students': students,
            'classrooms': Classroom.objects.all(),
            'total_students': students.count(),
            'current_classroom': classroom_id, # Untuk menandai pilihan di dropdown
        }
        return render(request, self.template_name, context)

@method_decorator(login_required, name='dispatch')
class StudentCreateView(View):
    template_name = 'students/form.html'

    def get(self, request):
        form = StudentForm()
        return render(request, self.template_name, {'form': form, 'title': 'Tambah Siswa'})

    def post(self, request):
        form = StudentForm(request.POST, request.FILES)
        if form.is_valid():
            student = form.save()
            messages.success(request, f'Siswa {student.full_name} berhasil ditambahkan.')
            return redirect('students:list')
        return render(request, self.template_name, {'form': form, 'title': 'Tambah Siswa'})

@method_decorator(login_required, name='dispatch')
class StudentUpdateView(View):
    template_name = 'students/form.html'

    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        form = StudentForm(instance=student)
        return render(request, self.template_name, {
            'form': form, 
            'title': 'Edit Siswa',
            'student': student
        })

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        form = StudentForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(request, f'Data {student.full_name} berhasil diperbarui.')
            return redirect('students:list')
        return render(request, self.template_name, {
            'form': form, 
            'title': 'Edit Siswa'
        })

@method_decorator(login_required, name='dispatch')
class StudentDeleteView(View):
    template_name = 'students/confirm_delete.html'

    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        return render(request, self.template_name, {'student': student})

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        nama_siswa = student.full_name
        student.delete()
        messages.success(request, f'Data siswa {nama_siswa} berhasil dihapus.')
        return redirect('students:list')

@method_decorator(login_required, name='dispatch')
class StudentDetailView(View):
    template_name = 'students/detail.html'
    
    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        return render(request, self.template_name, {'student': student})

@method_decorator(login_required, name='dispatch')
class StudentQRView(View):
    template_name = 'students/qr.html'
    
    def get(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        if not student.qr_code:
            student.generate_qr_code()
            student.save()
        return render(request, self.template_name, {'student': student})

@method_decorator(login_required, name='dispatch')
class StudentExportExcelView(View):
    def get(self, request):
        # 1. Ambil Parameter Filter dari URL
        classroom_id = request.GET.get('classroom')
        q = request.GET.get('q')

        # 2. Terapkan Filter yang sama dengan ListView
        students = Student.objects.select_related('classroom').all()
        
        if classroom_id:
            students = students.filter(classroom_id=classroom_id)
        if q:
            students = students.filter(
                Q(full_name__icontains=q) | Q(nis__icontains=q)
            )

        # 3. Proses Pembuatan Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Data Siswa'
        
        # Pengaturan Nama File jika ada filter kelas
        filename = "data_siswa_all.xlsx"
        if classroom_id:
            try:
                cls_name = Classroom.objects.get(id=classroom_id).name
                filename = f"data_siswa_kelas_{cls_name}.xlsx"
            except Classroom.DoesNotExist:
                pass

        # Headers
        headers = ['No', 'NIS', 'NISN', 'Nama Lengkap', 'Kelas', 'Telepon', 'Status']
        header_fill = PatternFill(start_color='1E40AF', end_color='1E40AF', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Menambahkan Data ke Baris
        for i, student in enumerate(students, 1):
            ws.append([
                i, 
                student.nis, 
                student.nisn, 
                student.full_name,
                str(student.classroom.name) if student.classroom else '-',
                student.phone if hasattr(student, 'phone') else '-',
                'Aktif' if student.is_active else 'Nonaktif',
            ])
            
        # 4. Return Response Download
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response